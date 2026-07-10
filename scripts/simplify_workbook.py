"""
simplify_workbook.py
Rebuilds Semptify_Master_Inventory_LIVE_reviewed.xlsx so it is
self-explanatory for anyone opening it cold:
  - Adds a "Start Here" sheet (plain-language, kid-simple guide)
  - Adds a "Files to Check First" column to Task Queue (for [JF] rows)
  - Color-codes Task Queue rows by tag ([RI]=blue, [EI]=green, [EF]=yellow, [JF]=red)
  - Translates status text to emoji versions
  - Inserts a legend row just below the header row
"""
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "Semptify_Master_Inventory_LIVE_reviewed.xlsx"

FONT_NAME = "Arial"

title_font = Font(name=FONT_NAME, bold=True, size=18, color="1F3864")
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
header_fill = PatternFill("solid", fgColor="1F3864")
body_font = Font(name=FONT_NAME, size=11)
big_body_font = Font(name=FONT_NAME, size=12)
wrap = Alignment(wrap_text=True, vertical="top")

STATUS_MAP = {
    "Done": "✅ Done",
    "In Progress": "🔄 Working On It",
    "Pending confirmation": "👀 Needs Review",
    "In Progress (branch, not merged)": "🔄 Working On It",
    "Screens 1-3 done, export ZIP pending (see TQ-003)": "🔄 Working On It",
    "Pending": "⭐ Not Started",
}

# Per-row file references for [JF] judgment-call tasks (keyed by data row index, 1-based after header)
# Update this dict as new JF tasks are added.
JF_FILE_REFS = {
    3: "app/core/product_manifest.py (see Module Inventory + Duplicates tabs)",
    6: "app/core/product_manifest.py, app/services/positronic_brain.py",
    7: "docs/Semptify_Site_GUI_Framework.md (RECORD pillar section), Duplicates tab",
}


def build_start_here(wb):
    if "Start Here" in wb.sheetnames:
        del wb["Start Here"]
    ws = wb.create_sheet("Start Here", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    lines = [
        ("title", "👋 Welcome! Here's How This Workbook Works"),
        ("space", ""),
        ("h", "What is this whole thing?"),
        ("body", "This workbook is like a to-do list for building Semptify. It tracks every job that needs"),
        ("body", "doing, who should do it, and whether it's done — so nothing gets lost or forgotten."),
        ("space", ""),
        ("h", "The tabs, in plain words:"),
        ("body", "📋  TASK QUEUE — the to-do list. Every job is one row."),
        ("body", "🤖  AI RESOURCE ROSTER — the helpers (AI tools) available, and what each one is good at."),
        ("body", "📦  MODULE INVENTORY — every piece of Semptify that exists and whether it works."),
        ("body", "🔍  GAP CHECK — a quick count-up so you don't have to count by hand."),
        ("space", ""),
        ("h", "The 4 job tags — every job gets ONE of these:"),
        ("body", "🔵  [RI]  =  Look something up. No building — just research. → Ask Gemini or MSN Copilot."),
        ("body", "🟢  [EI]  =  Build something small and simple. → Ask Windsurf (free) or GLM-5.2."),
        ("body", "🟡  [EF]  =  Build something big that touches many parts. → Ask Windsurf Premium."),
        ("body", "🔴  [JF]  =  Tricky decision. A person must think it through FIRST, then give it to Claude to build."),
        ("space", ""),
        ("h", "How to use the Task Queue (step by step):"),
        ("body", "1.  Look at the Status column. Find a row that says ⭐ Not Started or ⏸️ Stuck."),
        ("body", "2.  Look at the Tag. That tells you which helper to use (see above)."),
        ("body", "3.  If the row is 🔴 red ([JF]), check the 'Files to Check First' column — it tells you"),
        ("body", "    exactly which document(s) to read before making any decision."),
        ("body", "4.  Hand the job to the right helper. When done, change Status to ✅ Done."),
        ("space", ""),
        ("h", "Why the colors matter:"),
        ("body", "Each Task Queue row is colored by its tag so you can see the job type at a glance."),
        ("body", "Blue = research  |  Green = small build  |  Yellow = big build  |  Red = think first"),
        ("space", ""),
        ("h", "One rule to remember:"),
        ("body", "🔴  Red rows = STOP and think before doing anything. All other colors = just go build it."),
    ]

    for r, (kind, text) in enumerate(lines, start=1):
        cell = ws.cell(row=r, column=1, value=text)
        if kind == "title":
            cell.font = title_font
        elif kind == "h":
            cell.font = Font(name=FONT_NAME, bold=True, size=13, color="1F3864")
        elif kind == "body":
            cell.font = big_body_font
        cell.alignment = wrap


def simplify_task_queue(wb):
    if "Task Queue" not in wb.sheetnames:
        print("WARNING: 'Task Queue' sheet not found — skipping.")
        return
    ws = wb["Task Queue"]

    # --- Insert "Files to Check First" column at position 10 (before old column J) ---
    ws.insert_cols(10)
    ref_header = ws.cell(row=1, column=10)
    ref_header.value = "Files to Check First (🔴 red jobs only)"
    ref_header.font = header_font
    ref_header.fill = header_fill
    ref_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(10)].width = 38
    ws.row_dimensions[1].height = 30

    # Populate file refs (data starts at row 2; legend will be inserted at row 2 after this loop)
    for data_row_index, ref_text in JF_FILE_REFS.items():
        excel_row = data_row_index + 1  # +1 for header row
        cell = ws.cell(row=excel_row, column=10, value=ref_text)
        cell.font = body_font
        cell.alignment = wrap

    # --- Translate status values ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_cell = row[7]  # column H (0-indexed = index 7)
        if status_cell.value in STATUS_MAP:
            status_cell.value = STATUS_MAP[status_cell.value]

    # --- Color-code rows by tag ---
    blue_fill   = PatternFill("solid", fgColor="DDEBF7")
    green_fill  = PatternFill("solid", fgColor="E2EFDA")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    red_fill    = PatternFill("solid", fgColor="FCE4E4")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tag_val = str(row[1].value or "")  # column B
        if tag_val.startswith("[RI]"):
            fill = blue_fill
        elif tag_val.startswith("[EI]"):
            fill = green_fill
        elif tag_val.startswith("[EF]"):
            fill = yellow_fill
        elif "[JF]" in tag_val:
            fill = red_fill
        else:
            fill = None
        if fill:
            for cell in row:
                cell.fill = fill

    # --- Insert legend row just below the header ---
    ws.insert_rows(2)
    legend = {
        2: "🔵 [RI] = research",
        3: "🟢 [EI] = small build  |  🟡 [EF] = big build  |  🔴 [JF] = think-first decision",
    }
    for col_idx, text in legend.items():
        c = ws.cell(row=2, column=col_idx, value=text)
        c.font = Font(name=FONT_NAME, italic=True, size=9, color="555555")

    # --- Status dropdown validation ---
    dv = DataValidation(
        type="list",
        formula1='"⭐ Not Started,🔄 Working On It,⏸️ Stuck,👀 Needs Review,✅ Done,❌ Cancelled"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add("H3:H300")


def main():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found at: {WORKBOOK_PATH}")

    print(f"Loading: {WORKBOOK_PATH}")
    wb = load_workbook(str(WORKBOOK_PATH))
    print(f"Sheets before: {wb.sheetnames}")

    build_start_here(wb)
    simplify_task_queue(wb)

    wb.save(str(WORKBOOK_PATH))
    print(f"Sheets after:  {wb.sheetnames}")
    print("✅ Saved successfully.")


if __name__ == "__main__":
    main()
