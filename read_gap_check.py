import openpyxl

wb = openpyxl.load_workbook("Semptify_Master_Inventory_LIVE_reviewed.xlsx", data_only=True)

for sheet_name in ["Gap Check", "Stubs & TODOs", "Duplicates"]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ===")
        print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        print("Headers:", headers)
        for r in range(2, min(ws.max_row + 1, 20)):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            data = dict(zip(headers, row, strict=False))
            row_str = str(row).lower()
            if "capab" in row_str or "flag" in row_str or "gate" in row_str or "require_" in row_str:
                print(f"ROW {r}: {data}")
