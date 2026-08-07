import json

import openpyxl

wb = openpyxl.load_workbook('Semptify_Master_Inventory_LIVE_reviewed.xlsx', data_only=True)

print('Sheets:', wb.sheetnames)
print()

# --- Task Queue tab ---
if 'Task Queue' in wb.sheetnames:
    ws = wb['Task Queue']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    headers = [h for h in headers if h]  # strip trailing None columns
    print('Task Queue headers:', headers)
    print()
    tasks = []
    for r in range(2, ws.max_row+1):
        row = [ws.cell(r, c).value for c in range(1, len(headers)+1)]
        data = dict(zip(headers, row))
        # Skip fully empty rows
        if any(v for v in data.values() if v is not None):
            tasks.append(data)
            print(f"ROW {r}:")
            for h, v in data.items():
                if v is not None:
                    print(f"  {h}: {v}")
            print()
    # Also write JSON for the admin viewer
    with open('static/admin/task_queue_data.json', 'w', encoding='utf-8') as f:
        json.dump(tasks, f, indent=2, default=str)
    print(f"Written {len(tasks)} tasks to static/admin/task_queue_data.json")
else:
    print('Task Queue tab not found')

print()

# --- Module Inventory — capability/flag/gate rows only ---
if 'Module Inventory' in wb.sheetnames:
    ws = wb['Module Inventory']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    print('Module Inventory headers:', headers)
    print()
    for r in range(2, ws.max_row+1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        data = dict(zip(headers, row))
        row_str = str(row)
        if 'capab' in row_str.lower() or 'flag' in row_str.lower() or 'gate' in row_str.lower():
            print(f"ROW {r}:")
            for h, v in data.items():
                if v:
                    print(f"  {h}: {v}")
            print()
else:
    print('Module Inventory tab not found')
