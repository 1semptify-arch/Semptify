import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('Semptify_Master_Inventory_LIVE_reviewed.xlsx', data_only=True)
ws = wb['Task Queue']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]

print('Eligible rows (Tag=[EI] or [EF], Status=Open/Pending):')
print('-' * 60)
for r in range(2, ws.max_row+1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    data = dict(zip(headers, row))
    status = str(data.get('Status') or '')
    tag = str(data.get('Tag') or '')
    is_open = any(s in status.lower() for s in ['open', 'pending'])
    is_ei_ef = tag in ('[EI]', '[EF]')
    if is_open and is_ei_ef:
        print(f"ROW {r}: {data.get('Task ID')} | {tag} | {status}")
        print(f"  Description: {data.get('Description')}")
        print(f"  Pillar: {data.get('Pillar')}")
        print(f"  Target: {data.get('Target File/Path')}")
        print(f"  Verification: {data.get('Verification Step')}")
        print(f"  Depends On: {data.get('Depends On')}")
        print()
