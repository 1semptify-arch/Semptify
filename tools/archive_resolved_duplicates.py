#!/usr/bin/env python3
"""Archive resolved/rejected duplicate-resolve tasks out of the workbook.

THE ROOT CAUSE this fixes: workbook_bridge.py's rows_to_duplicate_tasks()
regenerates one task per row in the 'Duplicates' sheet on *every* sync,
unconditionally. Marking a task 'resolved' via mark_task_status.py correctly
sticks (merge_with_previous preserves it) — but the row itself never leaves
the sheet, so the task keeps being regenerated and keeps showing up in the
queue and dashboard, just wearing a green 'resolved' badge forever instead
of actually going away.

This script is the deliberate, human-run step that clears them out: for
every task marked 'resolved' or 'rejected', it removes the matching row
from the 'Duplicates' sheet and appends it to an 'Archive' sheet (created
if missing) — so history isn't lost, but the active queue only shows
active work. The next sync will stop regenerating a task for it entirely.

Usage:
    python tools/archive_resolved_duplicates.py            # do it
    python tools/archive_resolved_duplicates.py --dry-run  # preview only

Deliberately NOT part of the pre-commit hook or sync_orchestrator.py — this
is a decision to permanently move rows out of the human-maintained
'Duplicates' sheet, so a person runs it on purpose, not on every commit.
"""

import argparse
import hashlib
import json
import sys
from datetime import UTC, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

REPO_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = REPO_ROOT / "Semptify_Master_Inventory_LIVE_reviewed.xlsx"
TASKS_PATH = REPO_ROOT / "tools" / "agent_orchestrator_tasks.json"
ARCHIVABLE_STATUSES = {"resolved", "rejected"}


def stable_task_id(systems: str) -> str:
    """Must exactly match build_task()'s ID formula in workbook_bridge.py
    for duplicate rows, or matching will silently fail."""
    title = f"Resolve duplicate: {systems}"
    stable_key = f"app/core/product_manifest.py::{title}"
    return hashlib.sha1(stable_key.encode("utf-8")).hexdigest()[:12]  # noqa: S324 # nosec B324


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--dry-run", action="store_true", help="Show what would be archived without changing anything")
    args = parser.parse_args()

    if openpyxl is None:
        print("openpyxl is required for this script (it's in venv311).", file=sys.stderr)
        return 1
    if not WORKBOOK_PATH.exists():
        print(f"Workbook not found: {WORKBOOK_PATH}", file=sys.stderr)
        return 1
    if not TASKS_PATH.exists():
        print(f"Tasks file not found: {TASKS_PATH}", file=sys.stderr)
        return 1

    tasks = json.loads(TASKS_PATH.read_text(encoding="utf-8"))
    status_by_id = {t["id"]: t["status"] for t in tasks if t.get("category") == "duplicate_resolve"}

    try:
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
    except PermissionError:
        print(
            f"Could not open {WORKBOOK_PATH.name} — it's probably open in Excel. Close it and try again.",
            file=sys.stderr,
        )
        return 1

    if "Duplicates" not in wb.sheetnames:
        print("No 'Duplicates' sheet found.", file=sys.stderr)
        return 1
    ws = wb["Duplicates"]

    to_archive = []  # (row_index, row_values, resolved_status)
    for row_idx in range(2, ws.max_row + 1):
        row_values = [c.value for c in ws[row_idx]]
        if not any(row_values):
            continue
        systems = row_values[1] if len(row_values) > 1 else None
        if not systems:
            continue
        task_id = stable_task_id(systems)
        status = status_by_id.get(task_id)
        if status in ARCHIVABLE_STATUSES:
            to_archive.append((row_idx, row_values, status))

    if not to_archive:
        print("Nothing to archive — no resolved/rejected duplicate tasks match a current row.")
        return 0

    print(f"Found {len(to_archive)} row(s) to archive:")
    for _, row_values, status in to_archive:
        systems = row_values[1] if len(row_values) > 1 else "?"
        print(f"  [{status}] {systems}")

    if args.dry_run:
        print("\n--dry-run: no changes made.")
        return 0

    if "Archive" not in wb.sheetnames:
        archive_ws = wb.create_sheet("Archive")
        archive_ws.append(["ID", "Systems", "Overlap Description", "Details/Notes", "Archived Status", "Archived Date"])
    else:
        archive_ws = wb["Archive"]

    now = datetime.now(UTC).isoformat()
    for _, row_values, status in to_archive:
        archive_ws.append(list(row_values) + [status, now])

    # Delete from the bottom up so earlier row indices don't shift under us.
    for row_idx, _, _ in sorted(to_archive, key=lambda x: x[0], reverse=True):
        ws.delete_rows(row_idx, 1)

    wb.save(WORKBOOK_PATH)

    # Drop the same tasks from agent_orchestrator_tasks.json immediately,
    # rather than waiting for the next sync, so the GUI reflects it right away.
    archived_ids = {stable_task_id(row_values[1]) for _, row_values, _ in to_archive}
    remaining_tasks = [t for t in tasks if t["id"] not in archived_ids]
    TASKS_PATH.write_text(json.dumps(remaining_tasks, indent=2) + "\n", encoding="utf-8", newline="\n")

    print(f"\nArchived {len(to_archive)} row(s) to the 'Archive' sheet and removed them from the active queue.")
    print("Run sync_orchestrator.py next to refresh the embedded HTML views.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
